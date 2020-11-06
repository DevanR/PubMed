from datetime import datetime
import urllib.request
import re

from bs4 import BeautifulSoup

from Bio import Entrez
from openpyxl import load_workbook, Workbook

from writer import excel_writer, clear_output

EMAIL = 'your.email@example.com'
input_file = 'input/input.xlsx'


def search(query):
    Entrez.email = EMAIL
    handle = Entrez.esearch(db='pubmed',
                            sort='pub+date',
                            retmode='xml',
                            datetype='pdat',
                            mindate='2020/01/01',
                            maxdate=datetime.today().strftime('%Y/%m/%d'),
                            term=query)
    results = Entrez.read(handle)
    return results


def fetch_details(id_list):
    ids = ','.join(id_list)
    Entrez.email = EMAIL
    handle = Entrez.efetch(db='pubmed',
                           sort='pub+date',
                           retmode='xml',
                           id=ids,
                           datetype='pdat',
                           mindate='2020/01/01',
                           maxdate=datetime.today().strftime('%Y/%m/%d'))
    results = Entrez.read(handle)
    return results


def read_authors(file_name):
    wb = load_workbook(file_name)

    author_ws = wb['List 1_by names']
    author_list = []
    for i in range(1, author_ws.max_row + 1):
        if isinstance(author_ws.cell(row=i, column=1).value, int):
            author_list.append(
                author_ws.cell(row=i, column=4).value + '[Author]')

    return author_list


def read_institutes(file_name):
    wb = load_workbook(file_name)

    institute_ws = wb['List 2_by Institute']
    institute_list = []

    for i in range(1, institute_ws.max_row + 1):
        if isinstance(institute_ws.cell(row=i, column=1).value, int):
            institute_list.append(
                institute_ws.cell(row=i, column=2).value + '[Affiliation]')

    return institute_list

def parse_paper(paper):

    organisations = ['KK Women\'s and Children\'s Hospital', 'Kandang Kerbau Hospital']

    departments = []

    # Name of 1st author if from KKH
    first_author = "NA"
    if paper['MedlineCitation']['Article']['AuthorList'][0]['AffiliationInfo']:
        affiliation = paper['MedlineCitation']['Article']['AuthorList'][0]['AffiliationInfo'][0]['Affiliation']
        if any(org in affiliation for org in organisations):
            last_name = paper['MedlineCitation']['Article']['AuthorList'][0]['LastName']
            first_name = paper['MedlineCitation']['Article']['AuthorList'][0]['ForeName']
            first_author = "{}, {}".format(first_name, last_name)
            for org in organisations:
                if org in affiliation:
                    departments.append(org)

    # Name of last author if from KKH
    last_author = "NA"
    if paper['MedlineCitation']['Article']['AuthorList'][-1]['AffiliationInfo']:
        affiliation = paper['MedlineCitation']['Article']['AuthorList'][-1]['AffiliationInfo'][0]['Affiliation']
        if any(org in affiliation for org in organisations):
            last_name = paper['MedlineCitation']['Article']['AuthorList'][-1]['LastName']
            first_name = paper['MedlineCitation']['Article']['AuthorList'][-1]['ForeName']
            last_author = "{}, {}".format(first_name, last_name)
            for org in organisations:
                if org in affiliation:
                    departments.append(org)

    # KKH author if any
    author_list = []
    for author in paper['MedlineCitation']['Article']['AuthorList'][1:-1]:
        if author['AffiliationInfo']:
            affiliation = author['AffiliationInfo'][0]['Affiliation']
            if any(org in affiliation for org in organisations):
                any_last_name = author['LastName']
                any_first_name = author['ForeName']
                any_author = "{}, {}".format(any_first_name, any_last_name)
                author_list.append(any_author)
                for org in organisations:
                    if org in affiliation:
                        departments.append(org)

    if not author_list:
        author_list = "NA"
    else:
        author_list = ', '.join(author_list)

    # Dept
    if not departments:
        departments = "NA"
    else:
        departments = ', '.join(set(departments))

    # Authors
    authors = []
    for author in paper['MedlineCitation']['Article']['AuthorList']:
        if 'LastName' and 'ForeName' in author.keys():
            last_name = author['LastName']
            first_name = author['ForeName']
            author = "{}, {}".format(first_name, last_name)
        else:
            author = "{}".format(author['CollectiveName'])
        authors.append(author)

    # Name of Publication
    publication = paper['MedlineCitation']['Article']['ArticleTitle']

    # JournalInfo
    #title = paper['MedlineCitation']['Article']['Journal']['Title']
    ISO = paper['MedlineCitation']['Article']['Journal']['ISOAbbreviation']

    # PublishedDate
    pub_date = "NA"
    if paper['MedlineCitation']['Article']['ArticleDate']:
        year = paper['MedlineCitation']['Article']['ArticleDate'][0]['Year']
        month = paper['MedlineCitation']['Article']['ArticleDate'][0]['Month']
        day = paper['MedlineCitation']['Article']['ArticleDate'][0]['Day']
        pub_date = "Published {}/{}/{}".format(year, month, day)
    elif paper['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate']:

        if 'Year' in paper['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate'].keys():
            year = paper['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate']['Year']
        else:
            year =  ''
        if 'Month' in paper['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate'].keys():
            month = paper['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate']['Month']
        else:
            month =  ''
        if 'Day' in paper['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate'].keys():
            day = paper['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate']['Day']
        else:
            day =  ''
        pub_date = "Published {}/{}/{}".format(year, month, day)

    # DOI
    doi = 'NA'
    for id in paper['MedlineCitation']['Article']['ELocationID']:
        if id.attributes['EIdType'] == 'doi':
            doi = str(id)

    # IF
    impact_factor = 0
    try:
        ISSN = paper['MedlineCitation']['MedlineJournalInfo']['ISSNLinking'].replace('-', '')
        URL = "https://www.resurchify.com/impact-factor-search.php?query={}".format(ISSN)
        content = urllib.request.urlopen(URL)
        read_content = content.read()
        soup = BeautifulSoup(read_content,'html.parser')
        if soup.find_all(string=re.compile('IF:')):
            impact_factor = float(soup.find_all(string=re.compile('IF:'))[0].split(': ')[1])
    except Exception as e:
        print("Error: Unable to retrieve IF.")

    # PMID
    pmid = paper['MedlineCitation']['PMID']

    # IF=0
    if_zero = 1 if impact_factor == 0 else ''
    # IF<2
    if_less_than_two = 1 if  0 <impact_factor < 2 else ''
    # IF>2
    if_greater_than_two = 1 if impact_factor >=2 else ''

    country = paper['MedlineCitation']['MedlineJournalInfo']['Country']
    # National Journal
    national = 1 if country == 'Singapore' else ''
    # International Journal
    international = 1 if not(national) else ''

    row = {'first_author': first_author,
           'last_author': last_author,
           'authors': author_list,
           'dept': departments,
           'publication_name': ', '.join(authors) + '\n' + publication + '\n' + ISO + '\n' + pub_date + '\n' + doi,
           'if': impact_factor,
           'pmid': pmid,
           'if_zero': if_zero,
           'if_less': if_less_than_two,
           'if_more': if_greater_than_two,
           'national': national,
           'international': international}

    return row

def get_response(results):

    rows = []

    if results['IdList']:
        id_list = results['IdList']
        papers = fetch_details(id_list)

        for i, paper in enumerate(papers['PubmedArticle']):

            try:
                row = parse_paper(paper)
                row['S/N'] = i+1
                rows.append(row)
            except Exception as e:
                print(e)
    else:
        print('Nothing found!')

    return rows

if __name__ == '__main__':

    clear_output()

    # Get Authors
#    authors = read_authors(input_file)
#    author_results = search(' OR '.join(authors))
#    response = get_response(author_results)
#    excel_writer('Names', response)

    # Get Organisations
    institutes = read_institutes(input_file)
    org_results = search(' OR '.join(institutes))
    response = get_response(org_results)
    excel_writer('Institutes', response)
