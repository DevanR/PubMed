def clear_output():
    import os, re, os.path

    mypath = "output"
    for root, dirs, files in os.walk(mypath):
        for file in files:
            os.remove(os.path.join(root, file))


def excel_writer(name, rows):

    from openpyxl import load_workbook
    from shutil import copyfile
    from sys import exit, exc_info
    from copy import copy
    from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
    from openpyxl.styles import Alignment
    from datetime import date
    import os.path

    template = "templates/template.xlsx"

    # Copy and Create excel file if it does not exist
    output_file = "output/{}.xlsx".format(date.today())

    if not os.path.isfile(output_file):
        try:
            copyfile(template, output_file)
        except IOError as e:
            print("Unable to copy file. %s" % e)
            exit(1)
        except:
            print("Unexpected error:", exc_info())
            exit(1)

    wb = load_workbook(filename=output_file)
    ws = wb["{}".format(name)]

    # Update Inserted Row style
    offset = 2
    style_cell = "N1"
    for row in rows:
        # Insert empty rows
        ws.insert_rows(offset)
        # Apply style
        for row in ws.iter_cols(min_row=offset, max_row=offset, min_col=1, max_col=13):
            for cell in row:
                cell.style = copy(ws[style_cell].style)
                cell.font = copy(ws[style_cell].font)
                cell.border = copy(ws[style_cell].border)
                cell.fill = copy(ws[style_cell].fill)
                cell.number_format = copy(ws[style_cell].number_format)
                cell.protection = copy(ws[style_cell].protection)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Insert MCC data
    start_row = offset
    for i, row in enumerate(rows):
        ws.cell(row=start_row+i, column=1).value = row['S/N']
        ws.cell(row=start_row+i, column=2).value = row['first_author']
        ws.cell(row=start_row+i, column=3).value = row['last_author']
        ws.cell(row=start_row+i, column=4).value = row['authors']
        ws.cell(row=start_row+i, column=5).value = row['dept']
        ws.cell(row=start_row+i, column=6).value = row['publication_name']
        ws.cell(row=start_row+i, column=7).value = row['if']
        ws.cell(row=start_row+i, column=8).value = row['pmid']
        ws.cell(row=start_row+i, column=9).value = row['if_zero']
        ws.cell(row=start_row+i, column=10).value = row['if_less']
        ws.cell(row=start_row+i, column=11).value = row['if_more']
        ws.cell(row=start_row+i, column=12).value = row['national']
        ws.cell(row=start_row+i, column=13).value = row['international']

    # Add summation formulas to last row
    row_count = len(rows)
    sum_row = start_row + row_count
    ws.cell(row=sum_row, column=7).value = "=SUM(G2:G{})".format(row_count)
    ws.cell(row=sum_row, column=9).value = "=SUM(I2:I{})".format(row_count)
    ws.cell(row=sum_row, column=10).value = "=SUM(J2:J{})".format(row_count)
    ws.cell(row=sum_row, column=11).value = "=SUM(K2:K{})".format(row_count)
    ws.cell(row=sum_row, column=12).value = "=SUM(L2:L{})".format(row_count)
    ws.cell(row=sum_row, column=13).value = "=SUM(M2:M{})".format(row_count)

    # Adjust print area
    ws.print_area = "A1:M{}".format(sum_row)
    # Save file
    wb.save(filename=output_file)

